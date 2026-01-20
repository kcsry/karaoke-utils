import argparse
import html
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_ORDER = [
    "Anime",
    "Japani",
    "Korea",
    "Kiina",
    "My Little Pony",
    "Disney",
    "Englanti",
    "Suomi",
    "Muut",
]


def parse_sheet(ws):
    """Parse a worksheet into a list of (artist, title, source) tuples."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).lower() if h else "" for h in rows[0]]
    artist_idx = header.index("artist") if "artist" in header else None
    title_idx = header.index("title") if "title" in header else None
    source_idx = header.index("source") if "source" in header else None

    songs = []
    for row in rows[1:]:
        artist = row[artist_idx] if artist_idx is not None else None
        title = row[title_idx] if title_idx is not None else None
        source = row[source_idx] if source_idx is not None else None
        if title:  # Only include rows with a title
            songs.append((artist or "", title, source or ""))
    return songs


def get_sheets_to_process(wb, section_order):
    """Get ordered list of sheets to process."""
    sheets = []
    for name in section_order:
        if name in wb.sheetnames:
            sheets.append(name)
    for name in wb.sheetnames:
        if name not in sheets:
            sheets.append(name)
    return sheets


def group_songs_by_source(songs):
    """Group songs by source, sorted."""
    by_source = defaultdict(list)
    for artist, title, source in songs:
        by_source[source].append((artist, title))
    # Sort songs within each source by artist
    for source in by_source:
        by_source[source].sort(key=lambda s: s[0].lower())
    return sorted(by_source.items(), key=lambda p: p[0].lower())


# --- HTML Generation ---


def format_song_html(artist, title):
    """Format a song as HTML list item."""
    if artist:
        return f"<li>{html.escape(artist)} – {html.escape(title)}</li>"
    return f"<li>{html.escape(title)}</li>"


def generate_html(xlsx_path, section_order=None):
    """Generate HTML from XLSX file."""
    wb = load_workbook(xlsx_path)
    if section_order is None:
        section_order = DEFAULT_ORDER

    sheets_to_process = get_sheets_to_process(wb, section_order)
    html_parts = []

    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        songs = parse_sheet(ws)
        if not songs:
            continue

        html_parts.append(f"<h2>{html.escape(sheet_name)}</h2>")

        for source, source_songs in group_songs_by_source(songs):
            if source:
                html_parts.append(f"<h3>{html.escape(source)}</h3>")

            html_parts.append("<ul>")
            for artist, title in source_songs:
                html_parts.append(format_song_html(artist, title))
            html_parts.append("</ul>")

    return "\n".join(html_parts)


# --- Typst Generation ---


def escape_typst(text):
    """Escape special characters for Typst."""
    # Escape: # * _ ` $ @ \ < >
    text = re.sub(r"([#*_`$@\\<>])", r"\\\1", text)
    # Escape // (line comment) and /* (block comment)
    text = text.replace("//", "/\\/")
    text = text.replace("/*", "/\\*")
    return text


def format_song_typst(artist, title):
    """Format a song for Typst."""
    if artist:
        return f"- {escape_typst(artist)} – {escape_typst(title)}"
    return f"- {escape_typst(title)}"


TYPST_PREAMBLE = """\
#set page(paper: "a4", numbering: "1/1", margin: (x: 1.5cm, y: 2cm), header: context {
  let dominated-page = here().page()
  let dominated = query(heading.where(level: 1)).filter(h => h.location().page() <= dominated-page)
  if dominated.len() > 0 {
    let current = dominated.last()
    align(center, text(size: 16pt, weight: "bold", underline(current.body)))
  }
})
#set text(size: 8pt, font: "Lato")
#set par(leading: 0.5em)
#show heading: set block(sticky: true)
#show heading.where(level: 1): it => {}
#show heading.where(level: 2): set text(size: 10pt, weight: "bold")
#show heading.where(level: 2): set block(below: 0.7em)
"""


def generate_alphabetical_index(all_songs):
    """Generate an alphabetical index of all songs (yields lines)."""
    if not all_songs:
        return

    # Sort by title (case-insensitive)
    sorted_songs = sorted(all_songs, key=lambda s: s[0].lower())

    yield "#pagebreak()"
    yield ""
    yield "= Hakemisto"
    yield ""
    yield "#set text(size: 6pt)"
    yield "#set par(leading: 0.3em)"
    yield "#columns(3, gutter: 0.25cm)["
    yield "#table("
    yield "  columns: (1.5fr, 1fr),"
    yield "  stroke: none,"
    yield "  inset: 1.5pt,"
    yield "  row-gutter: 0pt,"
    current_letter = None
    last_title_artist = None
    for title, artist, section in sorted_songs:
        title_artist = str((title, artist)).lower()
        if title_artist == last_title_artist:
            continue  # Skip duplicates
        last_title_artist = title_artist

        # Get the first letter (uppercase) for grouping; non-alpha becomes "#"
        first_char = title[0].upper() if title else ""
        first_letter = first_char if first_char.isalpha() else "#"
        if first_letter != current_letter:
            current_letter = first_letter
            yield f"  table.cell(colspan: 2, align: center, inset: 3pt, fill: luma(80%))[#text(weight: \"bold\")[{escape_typst(current_letter)}]],"


        title_esc = escape_typst(title)
        artist_esc = escape_typst(artist) if artist else ""
        if artist_esc:
            yield f"  [{title_esc}], [{artist_esc}],"
        else:
            yield f"  [{title_esc}], [],"
    yield ")"
    yield "]"
    yield ""


def generate_typst(xlsx_path, section_order=None):
    """Generate Typst document from XLSX file (yields lines)."""
    wb = load_workbook(xlsx_path)
    if section_order is None:
        section_order = DEFAULT_ORDER

    sheets_to_process = get_sheets_to_process(wb, section_order)

    # Collect all songs for the alphabetical index
    all_songs = []  # List of (title, artist, section)

    yield TYPST_PREAMBLE

    first_section = True
    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        songs = parse_sheet(ws)
        if not songs:
            continue

        # Collect songs for the alphabetical index
        for artist, title, source in songs:
            all_songs.append((title, artist, sheet_name))

        # Page break before each section (except first)
        if first_section:
            first_section = False
        else:
            yield "#pagebreak()"
            yield ""

        # Full-width section heading
        yield f"= {escape_typst(sheet_name)}"
        yield ""

        use_columns = len(songs) >= 100
        if use_columns:
            yield "#columns(3, gutter: 1cm)["

        for source, source_songs in group_songs_by_source(songs):
            short_block = len(source_songs) < 30
            if short_block:  # emit unbreakable block
                yield "#block(breakable: false)["


            if source:
                yield f"== {escape_typst(source)}"
                yield ""

            for artist, title in source_songs:
                yield format_song_typst(artist, title)

            if short_block:
                yield "]"  # end of unbreakable block
            yield ""

        if use_columns:
            yield "]"
        yield ""

    # Generate alphabetical index
    yield from generate_alphabetical_index(all_songs)


# --- Main ---


def main():
    parser = argparse.ArgumentParser(description="Convert karaoke XLSX to documents")
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        default=Path("from-google-docs/Frostbite_2026_Karaoke.xlsx"),
        help="Input XLSX file",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output file (default: karaoke.html or karaoke.typ)",
    )
    parser.add_argument(
        "-f",
        "--format",
        choices=["html", "typst"],
        default="html",
        help="Output format (default: html)",
    )
    parser.add_argument(
        "--order",
        nargs="+",
        help="Section order (sheet names)",
    )
    args = parser.parse_args()

    section_order = args.order if args.order else DEFAULT_ORDER

    if args.format == "html":
        content = generate_html(args.input, section_order)
        output = args.output or Path("karaoke.html")
    else:
        content = "\n".join(generate_typst(args.input, section_order))
        output = args.output or Path("karaoke.typ")

    output.write_text(content, encoding="utf-8")
    print(f"Wrote {output}")


if __name__ == "__main__":
    main()
